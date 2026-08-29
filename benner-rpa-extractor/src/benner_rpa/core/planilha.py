"""G6 — a planilha original é imutável PARA O ROBÔ.

A distinção importa e custou um susto em 29/08/2026. O gate nasceu comparando o hash
com uma constante em `config.yaml`, o que enforça "este arquivo nunca muda". Errado: a
planilha é um documento de trabalho VIVO, e o responsável mantém à mão a coluna
`Benner OK`. Uma edição legítima dele derrubava o gate.

O que o gate precisa impedir é o ROBÔ escrever. Isso se verifica comparando o hash no
INÍCIO e no FIM da mesma execução — se mudou enquanto o robô rodava, foi ele.

A constante em config continua existindo, mas como registro informativo: uma mudança
entre execuções é REPORTADA, não fatal.
"""

from __future__ import annotations

import hashlib
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .estados import Estado
from .normalizacao import (
    eh_cnj_valido,
    nome_pasta_processo,
    normalizar_cabecalho,
    normalizar_processo,
)

NOME_CONTROLE = "planilha_controle.xlsx"

# Coluna de controle mantida À MÃO por uma pessoa. O robô LÊ e respeita; nunca escreve.
CANDIDATOS_COLUNA_BENNER = ["benner ok", "benner", "status benner"]

# Domínio informado pelo responsável em 29/08/2026.
BENNER_CONCLUIDO = 1        # baixado manualmente
BENNER_SO_PROCESSOS = 98    # só aparece em PROCESSOS (PASTAS); não há o que baixar
BENNER_NAO_ENCONTRADO = 99  # não existe no Benner

RESOLVIDOS_MANUALMENTE = {BENNER_CONCLUIDO, BENNER_SO_PROCESSOS, BENNER_NAO_ENCONTRADO}

COLUNAS_CONTROLE = [
    "processo_normalizado", "status", "pasta_benner", "docs_baixados",
    "docs_listados_popup", "docs_no_zip", "selecao_restantes_acionada",
    "arquivo_zip", "pedidos_exportados", "pasta_destino", "tentativas",
    "ultima_execucao", "observacao",
]


class PlanilhaAlterada(RuntimeError):
    """G6 violado: o hash do original mudou."""


@dataclass(frozen=True)
class LinhaProcesso:
    linha: int
    reclamada: str
    reclamante: str
    numero: str
    benner_ok: object = None      # marca humana; None = ainda por fazer
    obs: str = ""

    @property
    def resolvido_manualmente(self) -> bool:
        """Já tratado por uma pessoa — o robô não deve tocar."""
        return self.benner_ok in RESOLVIDOS_MANUALMENTE

    @property
    def marca_desconhecida(self) -> bool:
        """Marcado com algo fora do domínio combinado. Não é ignorável nem processável."""
        return self.benner_ok is not None and self.benner_ok not in RESOLVIDOS_MANUALMENTE

    @property
    def normalizado(self) -> str:
        return normalizar_processo(self.numero)

    @property
    def nome_pasta(self) -> str:
        return nome_pasta_processo(self.numero)


def sha256_planilha(caminho: Path) -> str:
    return hashlib.sha256(Path(caminho).read_bytes()).hexdigest()


def comparar_com_registro(caminho: Path, sha_registrado: str | None) -> tuple[bool, str]:
    """Compara com o hash anotado em config. Divergência é AVISO, não falha.

    O gate nasceu comparando com uma constante, o que enforçava "este arquivo nunca
    muda". Errado: a planilha é documento de trabalho vivo — o responsável mantém a
    coluna `Benner OK` à mão, e uma edição legítima dele derrubava o gate.

    O que o G6 precisa impedir é o ROBÔ escrever, e isso se verifica comparando o
    hash no início e no fim da MESMA execução (`conferir_integridade`). Entre
    execuções, mudança é esperada e só merece nota.
    """
    atual = sha256_planilha(caminho)

    if not sha_registrado:
        return True, f"sem hash registrado; atual = {atual}"
    if atual == sha_registrado:
        return True, "igual ao registrado"
    return True, (
        "a planilha mudou desde o ultimo registro — edicao manual e esperada.\n"
        f"        registrado: {sha_registrado}\n        atual:      {atual}"
    )


def conferir_integridade(caminho: Path, sha_esperado: str) -> None:
    """Levanta se o original mudou DURANTE a execução — aí foi o robô.

    Chamado com o hash lido no início: se divergir no fim, algo escreveu no original
    enquanto o robô rodava, e isso é o G6 violado.
    """
    atual = sha256_planilha(caminho)
    if atual != sha_esperado:
        raise PlanilhaAlterada(
            "G6: a planilha original foi alterada durante a execucao.\n"
            f"  no inicio: {sha_esperado}\n"
            f"  agora:     {atual}\n"
            "Algo ESCREVEU no original enquanto o robo rodava. Edicao humana entre "
            "execucoes e esperada e nao passa por aqui — ver `comparar_com_registro`."
        )


def detectar_coluna_processo(cabecalhos: list, candidatos: list[str]) -> int:
    """Índice 0-based da coluna do número do processo.

    Tolerante ao `º` de `Nº PROCESSO` — ver `normalizar_cabecalho`. Se nenhum
    candidato casar, cai para a primeira coluna cujos valores pareçam CNJ; se nem
    isso, levanta em vez de adivinhar.
    """
    normalizados = [normalizar_cabecalho(c) for c in cabecalhos]
    alvos = [normalizar_cabecalho(c) for c in candidatos]

    for alvo in alvos:
        if alvo in normalizados:
            return normalizados.index(alvo)

    for alvo in alvos:
        for i, n in enumerate(normalizados):
            if n and alvo in n:
                return i

    raise ValueError(
        f"coluna do processo nao encontrada. cabecalhos={cabecalhos} candidatos={candidatos}"
    )


def ler_processos(
    caminho: Path, aba: str, candidatos_coluna: list[str], *, sha_esperado: str | None = None
) -> list[LinhaProcesso]:
    """Lê o original em modo somente leitura. Nunca chama `save()`."""
    caminho = Path(caminho)
    if sha_esperado:
        conferir_integridade(caminho, sha_esperado)

    wb = load_workbook(caminho, read_only=True, data_only=True)
    try:
        if aba not in wb.sheetnames:
            raise ValueError(f"aba {aba!r} nao existe. abas={wb.sheetnames}")
        ws = wb[aba]
        linhas = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not linhas:
        return []

    cabecalho = list(linhas[0])
    idx = detectar_coluna_processo(cabecalho, candidatos_coluna)

    # Colunas opcionais mantidas à mão. Ausentes em planilhas antigas — por isso
    # `None` em vez de erro.
    normalizados = [normalizar_cabecalho(c) for c in cabecalho]
    idx_benner = next(
        (normalizados.index(a) for a in
         (normalizar_cabecalho(c) for c in CANDIDATOS_COLUNA_BENNER)
         if a in normalizados), None)
    idx_obs = normalizados.index("obs") if "obs" in normalizados else None

    def celula(linha, i):
        return linha[i] if i is not None and i < len(linha) else None

    processos: list[LinhaProcesso] = []
    for n, linha in enumerate(linhas[1:], start=2):
        bruto = linha[idx] if idx < len(linha) else None
        if bruto is None or not str(bruto).strip():
            continue
        processos.append(
            LinhaProcesso(
                linha=n,
                reclamada=str(linha[0] or "") if len(linha) > 0 else "",
                reclamante=str(linha[1] or "") if len(linha) > 1 else "",
                numero=str(bruto).strip(),
                benner_ok=celula(linha, idx_benner),
                obs=str(celula(linha, idx_obs) or ""),
            )
        )

    if sha_esperado:
        conferir_integridade(caminho, sha_esperado)

    return processos


def auditar_entrada(processos: list[LinhaProcesso]) -> dict:
    """Rede de segurança: duplicidade, formato e colisão de nome de pasta.

    Na planilha atual nada disto dispara (docs/GROUND-TRUTH.md). Roda mesmo assim,
    porque a próxima planilha é que vai precisar.
    """
    vistos: dict[str, list[int]] = {}
    pastas: dict[str, list[int]] = {}
    fora_do_formato: list[dict] = []

    for p in processos:
        vistos.setdefault(p.normalizado, []).append(p.linha)
        pastas.setdefault(p.nome_pasta.lower(), []).append(p.linha)
        if not eh_cnj_valido(p.numero):
            fora_do_formato.append({"linha": p.linha, "valor": p.numero})

    por_marca: dict[str, int] = {}
    for p in processos:
        chave = "pendente" if p.benner_ok is None else str(p.benner_ok)
        por_marca[chave] = por_marca.get(chave, 0) + 1

    return {
        "total": len(processos),
        "distintos": len(vistos),
        "por_marca_benner": por_marca,
        "pendentes": sum(1 for p in processos if p.benner_ok is None),
        "resolvidos_manualmente": sum(1 for p in processos if p.resolvido_manualmente),
        "marca_desconhecida": [
            {"linha": p.linha, "processo": p.numero, "valor": p.benner_ok}
            for p in processos if p.marca_desconhecida
        ],
        "duplicados": {k: v for k, v in vistos.items() if len(v) > 1},
        "colisoes_de_pasta": {k: v for k, v in pastas.items() if len(v) > 1},
        "fora_do_formato_cnj": fora_do_formato,
    }


def criar_controle(processos: list[LinhaProcesso], destino: Path) -> Path:
    """Cria a cópia de controle. Arquivo separado — o original nunca é tocado."""
    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Controle"
    ws.append(["RECLAMADA", "RECLAMANTE", "Nº PROCESSO", *COLUNAS_CONTROLE])

    for p in processos:
        ws.append([
            p.reclamada, p.reclamante, p.numero,
            p.normalizado, Estado.PENDENTE.value, "", "", "", "", "", "",
            "", p.nome_pasta, 0, "", "",
        ])

    wb.save(destino)
    wb.close()
    return destino
