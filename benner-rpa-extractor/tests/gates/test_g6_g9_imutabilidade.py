"""G6 — planilha original imutável.  G9 — pasta de referência somente leitura.

Ambos são gates de "nada mudou". A forma de provar é a mesma: hash antes, exercitar
o código, hash depois.
"""

import hashlib
from pathlib import Path

import pytest

from benner_rpa.core.planilha import (
    PlanilhaAlterada,
    auditar_entrada,
    conferir_integridade,
    criar_controle,
    detectar_coluna_processo,
    ler_processos,
    sha256_planilha,
)

pytestmark = pytest.mark.gate

RAIZ = Path(r"C:\MyWorkspace\claude-code\automacaoDeCastro")
PLANILHA = RAIZ / "data" / "CCR - Partes e Processos sem duplicidades.xlsx"
def _referencia() -> Path | None:
    """A pasta de referência é achada por padrão — seu nome traz um processo real
    e não pode ficar fixado num arquivo versionado."""
    achadas = sorted((RAIZ / "saida").glob("* - exemplo")) if (RAIZ / "saida").exists() else []
    return achadas[0] if achadas else None


REFERENCIA = _referencia()

# A planilha é documento VIVO: uma pessoa mantém a coluna `Benner OK` à mão. Fixar o
# hash num teste enforça "o arquivo nunca muda", que não é o que o G6 quer dizer.
# O gate real é início × fim da MESMA execução — ver `test_escrita_durante_execucao`.

CANDIDATOS = ["nº processo", "n processo", "numero_processo", "processo", "cnj"]


def _hashes(pasta: Path) -> dict[str, str]:
    return {
        p.name: hashlib.sha256(p.read_bytes()).hexdigest()
        for p in sorted(pasta.iterdir())
        if p.is_file()
    }


# ---------------------------------------------------------------- G6

def test_divergencia_com_o_registro_e_aviso_nao_falha():
    """Edição humana entre execuções é esperada e não pode derrubar o robô."""
    from benner_rpa.core.planilha import comparar_com_registro

    if not PLANILHA.exists():
        pytest.skip("planilha indisponivel neste ambiente")

    ok, msg = comparar_com_registro(PLANILHA, "0" * 64)
    assert ok, "hash diferente do registrado NAO pode ser fatal"
    assert "mudou desde o ultimo registro" in msg

    ok, msg = comparar_com_registro(PLANILHA, sha256_planilha(PLANILHA))
    assert ok and "igual" in msg


def test_escrita_durante_execucao_e_bloqueada(tmp_path):
    """O que o G6 realmente protege: o robô escrevendo no original.

    Simula uma execução — hash no início, algo escreve, conferência no fim.
    """
    original = tmp_path / "planilha.xlsx"
    original.write_bytes(b"conteudo original")
    sha_inicio = sha256_planilha(original)

    conferir_integridade(original, sha_inicio)          # nada mudou: passa

    original.write_bytes(b"alguem escreveu aqui")

    with pytest.raises(PlanilhaAlterada, match="durante a execucao"):
        conferir_integridade(original, sha_inicio)


def test_leitura_completa_nao_altera_o_original(tmp_path):
    if not PLANILHA.exists():
        pytest.skip("planilha indisponivel neste ambiente")

    antes = sha256_planilha(PLANILHA)

    processos = ler_processos(
        PLANILHA, "Partes e Processos", CANDIDATOS, sha_esperado=antes
    )
    auditar_entrada(processos)
    criar_controle(processos, tmp_path / "planilha_controle.xlsx")

    assert sha256_planilha(PLANILHA) == antes


def test_conferencia_detecta_alteracao(tmp_path):
    falsa = tmp_path / "p.xlsx"
    falsa.write_bytes(b"conteudo qualquer")

    with pytest.raises(PlanilhaAlterada):
        conferir_integridade(falsa, "0" * 64)


def test_controle_e_arquivo_separado(tmp_path):
    if not PLANILHA.exists():
        pytest.skip("planilha indisponivel neste ambiente")

    processos = ler_processos(PLANILHA, "Partes e Processos", CANDIDATOS)
    destino = criar_controle(processos, tmp_path / "planilha_controle.xlsx")

    assert destino.resolve() != PLANILHA.resolve()
    assert destino.exists()


def test_planilha_real_tem_333_processos_distintos():
    if not PLANILHA.exists():
        pytest.skip("planilha indisponivel neste ambiente")

    processos = ler_processos(PLANILHA, "Partes e Processos", CANDIDATOS)
    auditoria = auditar_entrada(processos)

    assert auditoria["total"] == 333
    assert auditoria["distintos"] == 333
    assert auditoria["duplicados"] == {}
    assert auditoria["colisoes_de_pasta"] == {}
    assert auditoria["fora_do_formato_cnj"] == []


def test_processo_da_pasta_de_referencia_esta_na_planilha():
    """A pasta de referência corresponde a uma linha real da planilha.

    O número é derivado do NOME DA PASTA, não fixado no teste: gravá-lo aqui
    publicaria um processo real num repositório.
    """
    if REFERENCIA is None or not PLANILHA.exists():
        pytest.skip("referencia ou planilha indisponivel neste ambiente")

    from benner_rpa.core.normalizacao import normalizar_processo

    alvo = normalizar_processo(REFERENCIA.name.replace(" - exemplo", ""))
    processos = ler_processos(PLANILHA, "Partes e Processos", CANDIDATOS)

    assert len([p for p in processos if p.normalizado == alvo]) == 1


def test_coluna_humana_e_lida_e_respeitada():
    """O robô lê a marca de quem baixou à mão e não refaz o trabalho."""
    if not PLANILHA.exists():
        pytest.skip("planilha indisponivel neste ambiente")

    processos = ler_processos(PLANILHA, "Partes e Processos", CANDIDATOS)
    a = auditar_entrada(processos)

    assert a["total"] == 333
    assert a["pendentes"] + a["resolvidos_manualmente"] + len(a["marca_desconhecida"]) == 333

    resolvidos = [p for p in processos if p.resolvido_manualmente]
    assert all(p.benner_ok in (1, 98, 99) for p in resolvidos)


def test_marca_fora_do_dominio_nao_e_ignorada():
    """Valor desconhecido não pode ser tratado como pendente nem como resolvido —
    seria adivinhar a intenção de quem escreveu."""
    if not PLANILHA.exists():
        pytest.skip("planilha indisponivel neste ambiente")

    processos = ler_processos(PLANILHA, "Partes e Processos", CANDIDATOS)
    desconhecidos = [p for p in processos if p.marca_desconhecida]

    for p in desconhecidos:
        assert not p.resolvido_manualmente
        assert p.benner_ok is not None


def test_deteccao_tolera_o_ordinal_do_cabecalho():
    """`Nº PROCESSO` é o cabeçalho real — o `º` não pode quebrar a detecção."""
    for variante in ["Nº PROCESSO", "N° PROCESSO", "nº processo", "Numero Processo", "  Nº  PROCESSO  "]:
        assert detectar_coluna_processo(["RECLAMADA", "RECLAMANTE", variante], CANDIDATOS) == 2


def test_deteccao_falha_em_vez_de_adivinhar():
    with pytest.raises(ValueError):
        detectar_coluna_processo(["COLUNA A", "COLUNA B"], CANDIDATOS)


# ---------------------------------------------------------------- G9

def test_pasta_de_referencia_intacta_apos_exercitar_o_codigo(tmp_path):
    if REFERENCIA is None:
        pytest.skip("pasta de referencia indisponivel neste ambiente")

    antes = _hashes(REFERENCIA)
    assert antes, "pasta de referencia vazia"

    # Exercita tudo que toca a referência: validação de ZIP e leitura do xlsx.
    from openpyxl import load_workbook

    from benner_rpa.core.zip_seguro import validar_zip

    for z in REFERENCIA.glob("*.zip"):
        validar_zip(z)
    wb = load_workbook(REFERENCIA / "Pedidos.xlsx", read_only=True, data_only=True)
    list(wb["Pedidos"].iter_rows(values_only=True))
    wb.close()

    assert _hashes(REFERENCIA) == antes


def test_saida_do_robo_nao_colide_com_a_referencia():
    """A pasta do robô é sem o sufixo ` - exemplo`. Nomes diferentes, por construção."""
    from benner_rpa.core.normalizacao import nome_pasta_processo

    nome_robo = nome_pasta_processo("0000000-22.2023.5.15.0002")

    assert nome_robo == "0000000-22.2023.5.15.0002"
    assert not nome_robo.endswith(" - exemplo")
    if REFERENCIA is not None:
        assert nome_robo != REFERENCIA.name
